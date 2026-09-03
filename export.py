"""Deliverables: the image for the client, the workbook for the file."""
from __future__ import annotations

import os

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import ChartConfig
from .engine import Prepared, breach_report, autocall_report, level_value

GREEN = "00915A"
NAVY = "0F2B46"


def export_images(fig, outdir: str, stem: str = "chart",
                  formats=("png", "pdf")) -> list[str]:
    """PNG for slides and email, PDF/SVG for print. 200 dpi is enough for a deck."""
    os.makedirs(outdir, exist_ok=True)
    paths = []
    for fmt in formats:
        p = os.path.join(outdir, f"{stem}.{fmt}")
        fig.savefig(p, dpi=200, facecolor="white")
        paths.append(p)
    return paths


def export_workbook(prep: Prepared, cfg: ChartConfig, path: str) -> str:
    """One workbook holding every number behind the picture.

    Sheets: Chart data (what is plotted) | Prices (aligned raw) | Observed
            (True = genuine quote, False = carried forward) | Quality | Barriers
            | Autocall | Config.
    """
    os.makedirs(os.path.dirname(os.path.abspath(path)) or ".", exist_ok=True)

    plotted = prep.plotted.copy()
    if cfg.show_worst_of and prep.plotted.shape[1] > 1:
        plotted[cfg.worst_of_label] = prep.worst_of
    for lv in cfg.levels:
        if lv.enabled:
            plotted[lv.label] = level_value(lv.pct, prep.base)

    config_rows = [
        ("Title", cfg.texts.title),
        ("Subtitle", cfg.texts.subtitle),
        ("Start", cfg.start), ("End", cfg.end),
        ("Initial fixing date", cfg.fixing_date),
        ("Scale", cfg.scale), ("Calendar", cfg.calendar),
        ("Base value", prep.base),
        ("Worst-of shown", cfg.show_worst_of),
        ("Source", cfg.texts.source),
        ("", ""),
        ("Levels", ""),
    ]
    for lv in cfg.levels:
        config_rows.append((f"  {lv.label}",
                            f"{lv.pct:.2%} of initial = {level_value(lv.pct, prep.base):,.4f}"
                            f"{'' if lv.enabled else '   (off)'}"))
    if cfg.autocall_dates:
        config_rows.append(("", ""))
        config_rows.append(("Autocall observation dates", ""))
        for d in sorted(cfg.autocall_dates):
            config_rows.append(("", d))
    if prep.warnings:
        config_rows.append(("", ""))
        config_rows.append(("Warnings", ""))
        for w in prep.warnings:
            config_rows.append(("", w))

    with pd.ExcelWriter(path, engine="openpyxl", datetime_format="dd-mmm-yy") as xl:
        plotted.to_excel(xl, sheet_name="Chart data")
        prep.prices.to_excel(xl, sheet_name="Prices")
        prep.observed.to_excel(xl, sheet_name="Observed")
        prep.quality.to_excel(xl, sheet_name="Quality", index=False)
        breach_report(prep, cfg).to_excel(xl, sheet_name="Barriers", index=False)
        ac = autocall_report(prep, cfg)
        if not ac.empty:
            ac.to_excel(xl, sheet_name="Autocall", index=False)
        pd.DataFrame(config_rows, columns=["Setting", "Value"]).to_excel(
            xl, sheet_name="Config", index=False)

        for name, ws in xl.book._sheets_dict.items() if hasattr(xl.book, "_sheets_dict") \
                else [(s.title, s) for s in xl.book.worksheets]:
            _format(ws)
    return path


def _format(ws):
    header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor=GREEN if ws.title == "Chart data" else NAVY)
    for cell in ws[1]:
        if cell.value is not None:
            cell.font = header
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "B2"
    widths = {}
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200)):
        for c in row:
            if c.value is not None:
                widths[c.column] = max(widths.get(c.column, 9),
                                       min(len(str(c.value)) + 2, 46))
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            if isinstance(c.value, float):
                c.number_format = "#,##0.00"
